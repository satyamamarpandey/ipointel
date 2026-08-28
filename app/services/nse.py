from __future__ import annotations
import re
from datetime import datetime, timezone
from io import BytesIO
import httpx
from bs4 import BeautifulSoup
from openpyxl import load_workbook

BASE="https://www.nseindia.com"
ARCHIVE_PAGE=f"{BASE}/static/regulations/segment-wise-historical-reports-capital-primary-market"

def _num(v):
    if v in (None,"","-","--"):return None
    if isinstance(v,(int,float)):return float(v)
    m=re.search(r"-?\d+(?:\.\d+)?",str(v).replace(",","").replace("₹",""));return float(m.group()) if m else None

def extract_list(payload):
    if isinstance(payload,list):return [x for x in payload if isinstance(x,dict)]
    if isinstance(payload,dict):
        for k in ("data","records","issues","result","results"):
            if isinstance(payload.get(k),list):return payload[k]
        out=[]
        for v in payload.values():
            if isinstance(v,list) and (not v or isinstance(v[0],dict)):out.extend(v)
        return out
    return []

def normalize(d:dict,status="Upcoming"):
    company=str(d.get("companyName") or d.get("company") or d.get("issuerName") or d.get("name") or d.get("symbol") or "Unknown IPO")
    symbol=str(d.get("symbol") or d.get("issueSymbol") or "")
    offered=_num(d.get("noOfSharesOffered") or d.get("sharesOffered") or d.get("issueSize"))
    bid=_num(d.get("noOfsharesBid") or d.get("noOfSharesBid") or d.get("sharesBid"))
    total=_num(d.get("noOfTime") or d.get("subscription") or d.get("totalSubscription"))
    if total is None and offered and bid: total=bid/offered
    def anynum(*keys):
        for k in keys:
            if d.get(k) not in (None, "", "-", "--"):
                return _num(d.get(k))
        return None
    return {
      "company":company,"symbol":symbol,"country":"India","exchange":"NSE/BSE","board":str(d.get("issueType") or d.get("series") or d.get("board") or "Mainboard"),
      "status":status,"sector":str(d.get("industry") or d.get("sector") or "Unknown"),"currency":"INR",
      "price_low":_num(d.get("issuePriceMin") or d.get("priceBandMin") or d.get("minPrice") or d.get("floorPrice")),
      "price_high":_num(d.get("issuePriceMax") or d.get("priceBandMax") or d.get("maxPrice") or d.get("capPrice") or d.get("issuePrice")),
      "lot_size":int(_num(d.get("marketLot") or d.get("lotSize") or d.get("minimumBidQuantity")) or 0) or None,
      "total_sub":total,"qib_sub":anynum("qibSubscription","qib","qibSub","qibNoOfTime"),"nii_sub":anynum("niiSubscription","hniSubscription","nii","niiSub","niiNoOfTime"),"retail_sub":anynum("retailSubscription","retail","retailSub","retailNoOfTime"),"open_date":str(d.get("issueStartDate") or d.get("openDate") or ""),"close_date":str(d.get("issueEndDate") or d.get("closeDate") or ""),
      "shares_offered_m":offered/1_000_000 if offered and offered>100_000 else offered,"raw":d
    }

def fetch_current():
    headers={"User-Agent":"Mozilla/5.0 IPOIntelligence/2.0","Accept":"application/json,text/plain,*/*","Referer":f"{BASE}/market-data/all-upcoming-issues-ipo"}
    rows=[]; warnings=[]; seen=set()
    with httpx.Client(headers=headers,timeout=20,follow_redirects=True) as c:
        try:c.get(BASE)
        except Exception:pass
        for label,url,status in [("current",f"{BASE}/api/ipo-current-issue","Open"),("upcoming",f"{BASE}/api/all-upcoming-issues?category=ipo","Upcoming")]:
            try:
                r=c.get(url);r.raise_for_status()
                for d in extract_list(r.json()):
                    x=normalize(d,status);key=(x["company"].lower(),x["symbol"].lower())
                    if key not in seen:seen.add(key);rows.append(x)
            except Exception as e:warnings.append(f"NSE {label}: {type(e).__name__}: {e}")
    return rows,warnings

def archive_links(html:str):
    soup=BeautifulSoup(html,"html.parser"); out=[]
    for a in soup.find_all("a",href=True):
        label=" ".join(a.stripped_strings)
        if "Primary Market Monthly Report" in label and ".xlsx" in (a.get("href") or ""):
            href=a["href"]
            if href.startswith("//"):href="https:"+href
            elif href.startswith("/"):href=BASE+href
            out.append((label,href))
    return out

def fetch_archive_links():
    with httpx.Client(headers={"User-Agent":"Mozilla/5.0 IPOIntelligence/2.0"},timeout=20,follow_redirects=True) as c:
        r=c.get(ARCHIVE_PAGE);r.raise_for_status();return archive_links(r.text)

def parse_monthly_xlsx(content:bytes):
    wb=load_workbook(BytesIO(content),data_only=True,read_only=True); rows=[]
    for ws in wb.worksheets:
        raw=list(ws.iter_rows(values_only=True))
        if not raw:continue
        header_idx=None; headers=[]
        for i,row in enumerate(raw[:25]):
            vals=[str(x).strip().lower() if x is not None else "" for x in row]
            joined=" | ".join(vals)
            if any(k in joined for k in ("company","issuer","issue name")) and any(k in joined for k in ("price","listing","issue")):
                header_idx=i;headers=vals;break
        if header_idx is None:continue
        for row in raw[header_idx+1:]:
            if not any(x not in (None,"") for x in row):continue
            d={headers[i]:row[i] for i in range(min(len(headers),len(row))) if headers[i]}
            company=next((str(v) for k,v in d.items() if v and any(t in k for t in ("company","issuer","issue name"))),"")
            if not company:continue
            def pick(*terms):
                for k,v in d.items():
                    if all(t in k for t in terms):return v
                return None
            symbol=next((str(v) for k,v in d.items() if v and "symbol" in k),"")
            listing_date=next((str(v) for k,v in d.items() if v and "listing" in k and "date" in k),"")
            rows.append({"company":company,"symbol":symbol,"listing_date":listing_date,"issue_price":_num(pick("issue","price")),"listing_price":_num(pick("listing","price")),"issue_size":_num(pick("issue","size")),"sheet":ws.title,"raw":{str(k):str(v) for k,v in d.items()}})
    return rows
